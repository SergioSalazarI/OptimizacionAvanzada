from json import load
from openpyxl import*
from gurobipy import*

# cargar parámetros
book = load_workbook("C:\\Users\\david\\OneDrive - Universidad de los Andes\\2022-2\\Optimización Avanzada\\Otros\\DatosExcursión.xlsx")
sheet = book.active

q = {} # PREFERENCIAS
for fila in range(3,15):
    for columna in range(3,7):
        excursioanista = sheet.cell(2,columna).value
        objeto = sheet.cell(fila,2).value
        q[(excursioanista,objeto)] = sheet.cell(fila,columna).value

p = {} # PESO
for fila in range(18,30):
    objeto = sheet.cell(fila,2).value
    p[objeto] = sheet.cell(fila,3).value

b = {} # CAPACIDAD
for fila in range(34,38):
    excursioanista = sheet.cell(fila,2).value
    b[excursioanista] = sheet.cell(fila,3).value

E = list(b.keys())
O = list(p.keys())

m = Model('Excusionistas')

x = m.addVars(E,O,vtype=GRB.BINARY,name='x')
#help(m.addVars)
for i in E:
    m.addConstr(quicksum(p[j]*x[i,j] for j in O)<=b[i])

for j in O:
    m.addConstr(quicksum(x[i,j] for i in E) ==1)

# FUNCIÓN OBJETIVO
m.setObjective(quicksum(q[i,j]*x[i,j] for i in E for j in O),GRB.MAXIMIZE)

m.update()
#m.optimize()

m.setParam("Outputflag",0)
m.optimize()

z = m.getObjective().getValue()
print(z)

for i,j in x.keys():
    if x[i,j].x>0:
        print("El objeto ",j," es llevado por: ",i)